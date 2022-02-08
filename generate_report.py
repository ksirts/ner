from matplotlib.font_manager import json_load
import pandas as pd
import wandb
import json
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter


class Report():
    def __init__(self, project):
        self.project = project
        self.run_result = self._load_run_result()

    def _load_run_result(self):
        api = wandb.Api()

        # Project is specified by <entity/project-name>
        runs = api.runs(self.project)

        summary_list, config_list, name_list = [], [], []
        for run in runs:
            # .summary contains the output keys/values for metrics like accuracy.
            #  We call ._json_dict to omit large files
            summary_list.append(run.summary._json_dict)

            # .config contains the hyperparameters.
            #  We remove special values that start with _.
            config_list.append(
                {k: v for k, v in run.config.items()
                 if not k.startswith('_')})

            # .name is the human-readable name of the run.
            name_list.append(run.name)

        runs_df = pd.DataFrame({
            "summary": summary_list,
            "config": config_list,
            "name": name_list
        })

        _JSON_FILE_NAME = 'project.json'
        runs_df.to_json(_JSON_FILE_NAME)

        return runs_df

    def _get_report(self, score_lst: list, ent_lst: list = None) -> list:
        data = json.loads(pd.DataFrame.to_json(self.run_result))
        report_lst = []
        if ent_lst is None:
            for i in range(len(data['name'])):
                reports = []
                sum = data['summary']
                model = sum[str(i)]
                reports.append(data['config'][str(i)]['seed'])
                reports.append(model[f'train/epoch'])
                for score in score_lst:
                    reports.append(model[f'eval/{score}'])
                report_lst.append(reports)
        else:
            for i in range(11):
                sum = data['summary']
                model = sum[str(i)]
                print(data['config'][str(i)]['seed'])
                for key, value in model.items():
                    if key.find(f'eval/TITLE') != -1:
                        print(key, value)

        return report_lst
    
    def _get_mean_values(self,report_lst:list)->list:
        row_title='Mean'
        mean_lst=[]
        mean_lst.append(row_title)
        
        

    def generate_excel_report(self):
        ent_lst = ['PER', 'GPE', 'LOC', 'ORG', 'PROD', 'EVENT',
                   'DATE', 'TIME', 'TITLE', 'MONEY', 'PERCENT']
        score_lst = ['f1', 'recall', 'precision']
        sum_title = ['model', '# of Epoch', 'f1', 'recall', 'precision']

        wb = Workbook()
        ws = wb.active
        ws.title = '00_Summary'
        ws.append(sum_title)
        
        # get reports
        report_lst =self._get_report(score_lst=score_lst)
        for report in report_lst:
            ws.append(report)

        wb.save('new_report.xlsx')


report = Report('d3auuv/huggingface')
report.generate_excel_report()
